import pandas as pd, numpy as np, re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df = pd.read_excel('/home/icarus/Desktop/scripts/data/KEA_UGCET2025_CutOff_Ranks.xlsx')
original_values = df.copy()

CATEGORIES = ['1G','1K','1R','2AG','2AK','2AR','2BG','2BK','2BR',
              '3AG','3AK','3AR','3BG','3BK','3BR','GM','GMK','GMP','GMR',
              'NRI','OPN','OTH','SCG','SCK','SCR','STG','STK','STR']

COLLEGE_CLEAN = {
    'Univesity of Visvesvaraya College of Engineering (A State Autonomous Public University on IIT Model) K R Circle, Bangalore': 'University Visvesvaraya College of Engineering',
    'Govt.S K S J T Institute of Engineering, Bangalore AMBEDKAR VEEDHI, K.R. CIRCLE, BANGALORE-01': 'Govt. S K S J T Institute of Engineering',
    'B M S College of Engineering, Basavanagudi, Bangalore (AUTONOMOUS) POST BOX NO 1908, BULL TEMPLE ROAD,BANGALORE': 'B M S College of Engineering',
    'Dr. Ambedkar Institute of Technology, Bangalore(AUTONOMOUS) OUTER RING ROAD,NEAR JNANA BHARATHI CAMPUS,MALLATHAHALLI,BANGALORE-560056': 'Dr. Ambedkar Institute of Technology',
    'R. V. College of Engineering, Bangalore(AUTONOMOUS) R.V. VIDYANIKETAN POST, MYSORE ROAD,BANGALORE': 'R. V. College of Engineering',
    'M S Ramaiah Institute of Technology, Bangalore(AUTONOMOUS) VIDYA SOUDHA,MSR NAGAR,MSRIT POST,BANGALORE - 560054.': 'M S Ramaiah Institute of Technology',
    'Dayananda Sagar College of Engineering, Bangalore(AUTONOMOUS) SHAVIGE MALLESHWARA HILLS, KUMARASWAMY LAYOUT, BANGALORE-560078': 'Dayananda Sagar College of Engineering',
    'Bangalore Institute of Technology, K.R.Road, Bangalore (AUTONOMOUS) K.R.ROAD, V.V.PURA, BANGALORE-560004': 'Bangalore Institute of Technology',
    'PES University 100 Feet Ring Road, Banashankari, 3rd Stage, Hosakerehalli, Near DSERT, , Bangalore KARNATAKA, pin code -560085': 'PES University',
    'M V J College of Engineering, Bangalore(AUTONOMOUS) NEAR ITPB,CHANNASANDRA,KADUGODI POST,BANGALORE': 'M V J College of Engineering',
    'New Horizon College of Engineering, Varthur, Bangalore(AUTONOMOUS) RING ROAD, NEAR MARATHALLI, BELLANDUR POST,KADUBHISANAHALLI,BANGALORE': 'New Horizon College of Engineering',
    'East West Institute Of Technology,Bangalore(AUTONOMOUS) SY. NO.63, OFF MAGADI MAIN ROAD, VISHWANEEDAM POST,BANGALORE': 'East West Institute of Technology',
    'Rajiv Gandhi Institute of Technology, Bangalore CHOLANGAR,R.T.NAGAR POST, HEBBAL,BANGALORE-560032': 'Rajiv Gandhi Institute of Technology',
    'Global Academy of Technology, Bangalore(AUTONOMOUS) IDEAL HOMES TOWNSHIP,RAJARAJESHWARINAGAR,BANGALORE-560098': 'Global Academy of Technology',
    'K S Institute of Technology, Raghuvanahalli, Bangalore(AUTONOMOUS) #14,RAGUVANAHALLI, KANAKAPURA MAIN ROAD,BANGALORE - 560 062.': 'K S Institute of Technology',
    'RNS Institute of Technology, Bangalore (AUTONOMOUS) CHANNASANDRA,UTTARAHALLI -KENGERI MAIN ROAD,SUBRAMANYAPURA POST,BANGALORE': 'RNS Institute of Technology',
    'Rajarajeswari College of Engineering, Bangalore (AUTONOMOUS) #14, RAMOHOLLI CROSS,KUMBALGODU, MYSORE ROAD,BANGALORE - 560074': 'Rajarajeswari College of Engineering',
    'S.J B. Institute of Technology, Bangalore (AUTONOMOUS) BGS HEALTH & EDUCATION CITY,NO 67, UTTARAHALLI ROAD.KENGERI, BANGALORE -60': 'S J B Institute of Technology',
    'Cambridge Institutute of Technology, K.R.Puram, Bangalore(AUTONOMOUS) K R PURAM,BANGALORE': 'Cambridge Institute of Technology',
    'Dayananda Sagar Academy of Technology & Management Technical Campus, Bangalore(AUTONOMOUS) UDAYAPURA,KANAKAPURA ROAD,BANGALORE URBAN': 'Dayananda Sagar Academy of Technology & Management',
    'DONBOSCO Institute of Technology, Bangalore(AUTONOMOUS) KUMBALAGODU, MYSORE ROAD, KENGERI HOBLI,BANGALORE-560074': 'Don Bosco Institute of Technology',
    'East Point College of Engineering & Technology, Bangalore(AUTONOMOUS) JNANA PRABHA, BIDARAHALLI,VIRGONAGAR POST,BANGALORE.': 'East Point College of Engineering & Technology',
    'R.R. Institute of Technology, Chikkabanavara, Bangalore(AUTONOMOUS) NO. 67, RAJA REDDY LAYOUT,NEAR CHIKKABANAVARA RAILWAY STATION,CHIKKABANAVARA,BANGALORE - 560090': 'R R Institute of Technology',
    'J S S Academy of Technical Education, Bangalore JSSATE CAMPUS,UTTARAHALLI-KENGERI ROAD ,BANGALORE 560060': 'J S S Academy of Technical Education',
    'AMC Engineering College, Bannerghatta Road, Bangalore(AUTONOMOUS) 18KM,BANNERAGHATTA ROAD,KALKERE,BANGALORE-560083,BANGALORE URBAN': 'AMC Engineering College',
    'Acharya Institute of Technology, Soldevanahalli,Chikkabanavara post, Bangalore SOLDEVANAHALLI, HESARAGHATTA MAIN ROAD, CHIKKABANAVARA POST,BANGALORE URBAN': 'Acharya Institute of Technology',
    "Atria Institute of Technology, Anand Nagar, Hebbal PO, Bangalore (AUTONOMOUS) ASKB CAMPUS, AG'S COLONY, 1ST MAIN, ANANDNAGAR,BANGALORE": 'Atria Institute of Technology',
    'Sri Krishna Institute of Technology, Bangalore NO. 29, HESARAGHATTA MAIN ROAD,CHIMNEY HILLS, CHIKKABANAVARA POST,BANGALORE - 560090': 'Sri Krishna Institute of Technology',
    'H.K.B.K.College of Engineering, Bangalore # 22/1, NAGAWARA, BANGALORE-560045': 'H K B K College of Engineering',
    'REVA University RUKMINI KNOWLEDGE PARK, KATTIGENAHALLI, YELAHANKA,BANGALORE - 560064,': 'REVA University',
    'Sambhram Institute of Technology, Bangalore AMBABHAVANI TEMPLE ROAD,M.S.PALYA, JALAHALLI EAST,VIDYARANYAPURA POST,BANGALORE': 'Sambhram Institute of Technology',
    'Sir M.Visvesvaraya Institute of Technology, Bangalore KRISHNADEVARAYANAGAR, HUNASAMARANAHALLI, VIA YELAHANKA,BANGALORE - 562 157': 'Sir M. Visvesvaraya Institute of Technology',
    'B M S Institute of Technology & Management, Yelahanka, Bangalore(AUTONOMOUS) POST BOX NO.6443, AVALAHALLI,DODDABALLAPURA MAIN ROAD,YELAHANKA, BANGALORE - 560064.': 'B M S Institute of Technology & Management',
    'M S Engineering College, Bangalore NAVARATHNA AGRAHARA, SADAHALLIPOST, NEAR BIAL, BANGALORE': 'M S Engineering College',
    'K N S Institute of Technology, Bangalore THIRUMENAHALLI, HEGGADENAGAR-KOGILU ROAD, YELAHANKA, BANGALORE-560064': 'K N S Institute of Technology',
    'C M R Institute of Technology, Kundala Halli village, Bangalore # 132, KUNDALAHALLI VILLAGE,IT PARK ROAD,BANGALORE - 560 037': 'C M R Institute of Technology',
    'Gopalan College of Engineering & Management, K.R. Puram, Bangalore SONNENAHALLI, HOODI VILLAGE, ITPL ROAD,BANGALORE': 'Gopalan College of Engineering & Management',
    'Impact College of Engineering & Applied Sciences, Bangalore KODIGEHALLI,SAHAKARANAGAR SOUTH BANGALORE': 'Impact College of Engineering & Applied Sciences',
    'T.John Institute of technology, Bangalore #86/1, KAMMANAHALLI,GOTTIGERE,BANNERUGATTA ROAD,BANGALORE': 'T John Institute of Technology',
    'Sai Vidya Institute of Technology, Rajanakunte, Bangalore RAJANUKUNTE,DODDBALLAPURA ROAD ,VIA: YELAHANKA,BANGALORE': 'Sai Vidya Institute of Technology',
    'Vijaya Vittala Institute of Technology, Doddagubbi, Hennur Bagalur Road, Bangalore #35/1, DODDA GUBBI POST, HENNUR-BAGALURU ROAD,BANGALORE': 'Vijaya Vittala Institute of Technology',
    'Vemana Institute of Technology, Bangalore #1, MAHAYOGI VEMANA ROAD, 3RD BLOCK, KORMANGALA, BANGALORE -560034.': 'Vemana Institute of Technology',
    'Jyothi Institute of Technology, Kanakapura Road,Bangalore ,BANGALORE': 'Jyothi Institute of Technology',
    'East West College Of Engineering, Yelahanka,Bangalore NO.13 SECTOR "A", YELAHANKA NEW TOWN,YELAHANKA, BANGALORE-560064': 'East West College of Engineering',
    'BGS College of Engineering & Technology,Bangalore CA SITE NO. 6 & 7, 3RD MAIN, PIPE LINE ROAD, MAHALAKSHMIPURAM, WEST OF CHORD ROAD': 'BGS College of Engineering & Technology',
    'Brindavan College of Engineering, Yelahanaka, Bangalore DWARAKANAGARA, BAGALUR MAIN ROAD,YELAHANKA, BANGALORE - 560 063': 'Brindavan College of Engineering',
    'S E A College of Engineering & Technology, Virgonagar, Bangalore EKTA NAGAR, BASAVANAPURA, VIRGO NAGAR POST, BANGALORE - 560 049.': 'S E A College of Engineering & Technology',
    'APS College of Engineering, Somanahalli, Bangalore 26 KM KANAKAPURA ROAD BANGALORE 82,BANGALORE URBAN': 'APS College of Engineering',
    'K S School of Engineering & Management, Mallasandra, Bangalore #15/1, MALLASANDRA, OFF KANAKAPURA ROAD, BANGALORE - 560062': 'K S School of Engineering & Management',
    'Sri Revana Siddeswara Institute of Technology, Bangalore CHOKKANAHALLI, CHIKKAJALA,BANGALORE NORTH - 562157': 'Sri Revana Siddeswara Institute of Technology',
    'R V Institute Of Technology and Management, Bengaluru KOTHANUR, 8TH PHASE J.P.NAGAR, BENGALURU-560 078.': 'R V Institute of Technology and Management',
    'B N M Institute of Technology, Bangalore (AUTONOMOUS) 12TH MAIN, 27TH CROSS, BSK SECOND STAGE BANGALORE 560070': 'B N M Institute of Technology',
    'Ghousia Institute of Technology for Women, Bangalore Hosur Road, DRC Post, Banglaore-29': 'Ghousia Institute of Technology for Women',
    'Sri Sairam College Of Engineering, Anekal, Bangalore SAI LEO NAGAR GUDDANAHALLI VILLAGE,SAMANDUR POST,ANEKAL': 'Sri Sairam College of Engineering',
    'Bangalore Technological Institute, Sarjapura Road, Bangalore SY NO.66/6,66/5,67/6,&67/5, KODATHI VILLAGE, VARTHOOR HOBLI,BANGALORE EAST TQ BANGALORE URBAN': 'Bangalore Technological Institute',
    'PES UNIVERSITY(Electronic City Campus) Hosur Rd, Konappana Agrahara, Electronic City, Bengaluru, Karnataka 560100': 'PES University (Electronic City Campus)',
    'ACS College of Engineering, Mysore Road, Bangalore 207,KAMBIPURA, MYSORE ROAD, KENGERI HOBLI,BANGALORE - 560074.': 'ACS College of Engineering',
    'R.L.Jalappa Institute of Technology, Doddaballapura KODIGEHALLI, DODDABALLAPUR-561203, BANGALORE RURAL': 'R L Jalappa Institute of Technology',
    'Cambridge Institute Of Technology, North Campus, Devanahalli, Bangalore SURVEY NO. 73, LINGADHEERA GOLLAHALLI, KUNDANA POST, DEVANAHALLI TALUK': 'Cambridge Institute of Technology (North Campus)',
    'Nagarjuna College of Engineering & Technology,venkatagiri Kote, Devanahalli, Bangalore Rural.(AUTONOMOUS) MUDUGURKIVENKATAGIRI KOTE DEVANAHALLI,BANGALORE': 'Nagarjuna College of Engineering & Technology',
    'Dr. Sri Sri Sri Shivakumara Mahaswamyji College of Engineering, Byranayakanahalli, Bangalore Rural BYARANAYAKANAHALLI,DODDABALE POST,NELAMANGALA TALUK,BANGALORE RURAL DISTRICT': 'Dr. Sri Sri Sri Shivakumara Mahaswamyji College of Engineering',
    'Sir M V School of Architecture,Hunasamaranahalli, Bangalore KRISHNADEVARAYANAGAR,HUNASAMARANAHALLI,VIA YELAHANKA,BANGALORE - 562 157': 'Sir M V School of Architecture',
    'M . S . Ramaiah University of Applied Sciences UNIVERSITY HOUSE, GNANAGANGOTHRI CAMPUS, NEW BEL ROAD, MSR ROAD, BANGALORE - 560054': 'M S Ramaiah University of Applied Sciences',
    'ALLIANCE University Central Campus, Chikkahadage Cross Chandapura-Anekal, Main Road, Bengaluru, Karnataka 562106': 'Alliance University',
    'CMR University Main Campus, Off Hennur Bagalur Main Road, Chagalatti, Bangalore': 'CMR University',
    'RV University R.V. VIDYANIKETAN POST, MYSORE ROAD,BANGALORE': 'RV University',
    'Garden City University 16th KM, Old Madras Road (Near K R Puram), Bangalore, Karnataka, 560049.': 'Garden City University',
    'AMITY UNIVERSITY National Highway 648 (old 207), Devanahalli - Doddaballapur Road': 'Amity University',
    'Dayananda Sagar University DEVARAKAGGALAHALLI VILLAGE, HAROHALLI, KANAKAPURA MAIN ROAD, RAMANAGAR DIST': 'Dayananda Sagar University',
    'Vidyashilp University #125, Bettenahalli, Kundana Hobli, Chapparkallu Rd, Bengaluru, Karnataka 562110': 'Vidyashilp University',
    'SAPTHAGIRI NPS UNIVERISTY #14/5, CHIKKASANDRA,HESARAGHATTA MAIN ROAD,BANGALURU': 'Sapthagiri NPS University',
    'PRESIDENCY University ITGALPURA, RAJANUKUNTE, YELAHANKA, BENGALURU - 560 064': 'Presidency University',
    'Sampoorna Group of Institutions, Channapatana, Ramanagara District BELEKERE, MALLUR HOBLI, CHANNAPATNA TALUK, RAMANAGAR,DISTRICT': 'Sampoorna Group of Institutions',
    'P A College of Engineering, Kairangal, Bantwala Tq,. Mangalore NADUPADAV, NEAR MANGALORE UNIVERSITY,MONTEPADAV POST, KAIRANGALA VILLAGE,BANTWAL-MANGALORE': 'P A College of Engineering',
    'A J Institute Of Engineering And Technology.Kottar chowki Boloor Village Mangalore KOTTARA CHOWKI BOLOOR VILLAGE,DAKSHINA KANNADA': 'A J Institute of Engineering and Technology',
    'St.Joseph Engineering College, Mangalore(AUTONOMOUS) MANGALORE': 'St. Joseph Engineering College',
    "Sahyadri College Of Engineering & Management, Adyar, Mangalore(AUTONOMOUS) 'SAHYADRI CAMPUS'N.H-48, ADYAR,MANGALORE": 'Sahyadri College of Engineering & Management',
    'Mangalore Institute of Technology & Engineering, Moodabidri, Mangalore(AUTONOMOUS) BADAGA MIJAR, MOODBIDRI,MANGALORE TQ-574225,DK': 'Mangalore Institute of Technology & Engineering',
    "Alva's Institute of Engineering & Technology, Moodabidre, D.K (AUTONOMOUS) SHOBHAVANA CAMPUS,MIJAR,MOODBIDRI - 574225,MANGALORE TALUK,DHAKSHINA KANNADA": "Alva's Institute of Engineering & Technology",
    'Srinivas Institute of Technology, Mangalore SRINIVAS CAMPUS VALACHIL, ARKULA VILLAGE,FARANGIPETE POST,MANGALORE': 'Srinivas Institute of Technology',
    'Karavali Institute of Technology, Neermarga, Mangalore PADU POST, VIA PEDUMALE,NEERUMARGA,MANGALORE -575029': 'Karavali Institute of Technology',
    'Shreedevi Institute of Technology, Mangalore AIRPORT ROAD, KENJAR VILLAGE,MALAVOOR PANCHAYAT,DAKSHINA KANNADA': 'Shreedevi Institute of Technology',
    'Yenepoya Institute Of Technology, Mangalore VIDYANAGAR N.H 13THODAR MIJAR POST MOODBIDRI MANGALORE TQ': 'Yenepoya Institute of Technology',
    "Beary's Institute of Technology, Boliar Village,Bantwal Tq, Mangalore LANDS END, INNOLI,BOLIYAR VILLAGE,NEAR MANGALORE UNIVERSITY,MANGALORE-574 153": "Beary's Institute of Technology",
    'Canara Engineering College Bantwal, BENJANAPADAVU,BANTWAL,PIN 574219,DAKSHINA KANNADA DIST,': 'Canara Engineering College',
    'S D M Institute of Tech., Ujire, Dakshina Kannada UJIRE-574240 DAKSHINA KANNADA KARNATAKA': 'S D M Institute of Technology',
    'K V G College of Engineering, Sullia KURUNJIBHAG, SULLIA TALUK,DAKSHINA KANNADA - 574327': 'K V G College of Engineering',
    'Moodalakatte Institute of Technology, Kundapura, Udupi MOODLAKATTE,KUNDAPURA TALUK,UDUPI DISTRICT': 'Moodalakatte Institute of Technology',
    'Shri Madhwa Vadiraja Institute of Technology and Management, Udupi VISHWOTHAMA NAGAR,': 'Shri Madhwa Vadiraja Institute of Technology and Management',
    'Vivekananada College of Engineering Technology, Puttur NEHARU NAGARA,PUTTUR TALUK,DAKSHINA KANNADA': 'Vivekananda College of Engineering Technology',
    'Srinivas University SRINIVAS CAMPUS MUKKA MANGALURU': 'Srinivas University',
    'JSS Science and Technology University JSS TECHNICAL INSTITUTIONS CAMPUS, MANASAGANGOTHRI, MYSURU 570006': 'JSS Science and Technology University',
    'Sri Jayachamarajendra College of Engineering(Constituent College of JSS Science & Technology University), Mysore JSS TECHNICAL INSTITUTIONS CAMPUS,MANASAGANGOTHRI PO,MYSORE': 'Sri Jayachamarajendra College of Engineering',
    'The National Institute of Engineering(SOUTH CAMPUS), Mysore(AUTONOMOUS) MANANDAVADI ROAD,MYSORE 570008': 'The National Institute of Engineering (South Campus)',
    'The National Institute of Engineering(NORTH CAMPUS), No.50(Part) Koorgalli,Hootgalli Industrial Area ,Mysuru-570018': 'The National Institute of Engineering (North Campus)',
    'Vidya Vardhaka College of Engineering, Mysore(AUTONOMOUS) P.B.NO.206,GOKULAM 3RD STAGE,MYSORE': 'Vidya Vardhaka College of Engineering',
    'P E S College of Engineering, Mandya(AUTONOMOUS) MANDYA: 571 401': 'P E S College of Engineering',
    'Maharaja Institute of Technology Mysore,Belawadi,Srirangapatna,Mandya(AUTONOMOUS) BEHIND K R MILLS,BELAWADI VILLAGE,NAGUVANAHALLI PANCHAYAT,SRIRANGAPATNA TALUK,MANDYA DISTRICT': 'Maharaja Institute of Technology Mysore',
    'Maharaja Institute of Technology,Tandavapura,Mysore THANDAVAPURA VILLAGA,ON NATIONAL HIGHWWAYNANJANAGUD TALUK,MYSORE DISTRICT': 'Maharaja Institute of Technology Tandavapura',
    'Mysore College Of Engineering and Management,Mysore CHIKAHALLI, VARUNA HOBALI, NEAR BIG BANYAN TREE, TN ROAD,MYSORE': 'Mysore College of Engineering and Management',
    'ATME College of Engineering, Mysore SY.NO.1002, HAROHALLI (MELLAHALLI),MYSORE': 'ATME College of Engineering',
    'Vidya Vikas Institute of Engineering & Technology, Mysore #127,ALANAHALLY,ALANAHALLY POST,MYSORE-BANNUR ROAD,MYSORE': 'Vidya Vikas Institute of Engineering & Technology',
    'Mysuru Royal Institute Of Technology,Lakshmipura road,Mysuru LAKSHMIPURA ROAD, OFF MYSURU -BENGALURU HIGHWAY,MANDYA': 'Mysuru Royal Institute of Technology',
    'G S S S Institute of Engineering & Technology for Women, Mysore K.R.S. ROAD, METAGALLI INDUSTRIAL AREA, MYSORE-16': 'G S S S Institute of Engineering & Technology for Women',
    'School of Planning and Architchure, University of Mysore 2nd Floor, Senate Bhavan, Manasagangotri P.O., Mysore': 'School of Planning and Architecture, University of Mysore',
    'CAUVERY COLLEGE OF ENGINEERING KBL LAYOUT, ALANAHALLY MYSORE': 'Cauvery College of Engineering',
    'Cauvery institute of Technology, Koppalu gate, Sundahally, Mandya SUNDAHALLI, SIDDAIAHNA KOPPALU GATE, YELIHUR POST, KOTHATHI HOBLI, MANDYA': 'Cauvery Institute of Technology',
    'Adhichunchanagiri University (Formerly B G S Institute of Technology) B.G.NAGARA ,NAGAMANGALA TALUK,MANDYA DISTRICT -571448': 'Adhichunchanagiri University',
    'P E S Institute of Technology & Management, Shimoga GUDDADAARAKERE, KOTEGANGOORSAGAR ROAD,SHIMOGA': 'P E S Institute of Technology & Management',
    'Jawaharlal Nehru New College of Engineering, Shimoga P.B. NO. 128,NAVULE, SHIMOGA': 'Jawaharlal Nehru New College of Engineering',
    'Malnad College of Engineering, Hassan(AUTONOMOUS) P.B. NO. 21, SALAGAME ROAD,HASSSAN': 'Malnad College of Engineering',
    'Government Engineering College, Hassan DAIRY CIRCLE, B M ROAD, HASSAN': 'Government Engineering College, Hassan',
    'Government Engineering College, Mosale Hosahalli, Hassan MOSALEHOSAHALLI, HASSAN': 'Government Engineering College, Mosale Hosahalli',
    'NAVKIS COLLEGE OF ENGINEERING, KIADB Industrial Area,Thimmanahally, Hassan P.B.NO. 55, NH-48, KIADB INDUSTRIAL AREA, THIMMANAHALLY, HASSAN-573201': 'NAVKIS College of Engineering',
    'Rajeev Institute of Technology, Hassan PLOT 1-D(P-1), GROWTH CENTRE, INDUSTRIAL AREA, BANGALORE MANGALORE BYPASS ROAD HASSAN-573201': 'Rajeev Institute of Technology',
    'Bahubali College of Engineering, Shravanabelagola, Hassan GOMMATANAGARA,SHRAVANABELAGOLA - 573135 HASSAN': 'Bahubali College of Engineering',
    'Bapuji Institute of Engineering & Technology, Davangere (AUTONOMOUS) BAPUJI INSTITUTE OF ENGINEERING & TECHNOLOGY,P.B. NO.325,SHAMANUR ROAD,DAVANGERE-577 004': 'Bapuji Institute of Engineering & Technology',
    'University B.D.T College of Engineering, Davanagere P J EXTENSION,HADADI ROAD,DAVENGERE': 'University B D T College of Engineering',
    'University B.D.T. college of Engineering, (H.GOV),Davangere': 'University B D T College of Engineering (H. Gov)',
    'G M Institute of Technology, Davanagere POST BOX. NO. 4, P.B.ROAD.DAVANGERE.': 'G M Institute of Technology',
    'Jain Institute of Technology, Davanagere 323, BADA CROSS NEAR BLIND SCHOOL AVARAGERE DAVANGERE': 'Jain Institute of Technology',
    'GM University KARUR VILLAGE, KASABA HOBLI, DAVANGERE TALUK AND DISTRICT': 'GM University',
    'G Madegowda Institute of Technology, Bharathinagara, Maddur, Mandya BHARATHINAGARA (K M DODDY) MADDUR TQ,MANDYA': 'G Madegowda Institute of Technology',
    'Jnanavikasa Institute of Technology, Bidadi,Ramanagar VIDYA SOUDHA, 32ND K.M.BANGALORE - MYSORE HIGH WAY,': 'Jnanavikasa Institute of Technology',
    'Ghousia Engineering College, Ramanagara BANGALORE - MYSORE ROAD,RAMANAGARA': 'Ghousia Engineering College',
    'Amrutha Institute of Engineering & Management Science (AIEMS), Ramanagar BIDADI INDUSTRIAL AREA, OFF MYSORE ROAD, NEAR TOYOTA KIRLOSKAR MOTORS, BMRDA REGION, BANGALORE': 'Amrutha Institute of Engineering & Management Science',
    'KLE Technological University(Formerly (BVBCET) BVBHOOMARADDI COLLEGE CAMPUS, VIDYANAGAR, HUBBALLI': 'KLE Technological University',
    'K L E Technological Univeristy, Belgaum Campus (Formerly KLE Dr M.S.Sheshagiri College of Engineering and Technology) UDYAMBAGH,BELGAUM': 'K L E Technological University (Belgaum Campus)',
    'K.L.S. Gogte Institute of Technology, Belgaum.(AUTONOMOUS) "JNANA GANGA", UDYAMBAG,BELGAUM': 'K L S Gogte Institute of Technology',
    'SDM College of Engineering, Dharwad (AUTONOMOUS) DHAVALAGIRI,KALGHATGI ROAD,DHARWAD': 'SDM College of Engineering',
    'Basavakalyana Engineering College, Basavakalyana, Bidar District N.H.9 ,BASAVAKALYAN-585327,DIST. BIDAR': 'Basavakalyana Engineering College',
    'P D A College of Engineering, Gulbarga(AUTONOMOUS) AIWAN-E-SHAHI AREA,STATION ROAD,GULBARGA-585102': 'P D A College of Engineering',
    'Navodaya Institute of Technology, Raichur(AUTONOMOUS) NAVODAYA NAGAR, MANATRALAYAM ROAD, RAICHUR': 'Navodaya Institute of Technology',
    "H K E's Society's Sir M Visvesvaraya College of Engineering, Raichur YERAMARUS CAMP,RAICHUR-584 135,": "H K E Society's Sir M Visvesvaraya College of Engineering",
    'Rao Bahadur Y.Mahabaleswarappa Engineering College, Bellary CONTONMENT, BELLARY-583104': 'Rao Bahadur Y. Mahabaleswarappa Engineering College',
    'Proudadevaraya Institute of Technology, Hospet SHAH BHAVARLAL BABULAL NAHAR CAMPUS,T.B.DAM ROAD, BESIDE TSP LIMITED,HOSPET - 583225,BELLARY DISTRICT': 'Proudadevaraya Institute of Technology',
    'Ballari Institute of Technology & Management, Bellary(AUTONOMOUS) "JNANA GANGOTRI" CAMPUS, #873/2 OF KOLAGAL VILLAGE,BELLARY-HOSPET ROAD, ALLIPUR': 'Ballari Institute of Technology & Management',
    'KISHKINDA UNIVERSITY BALLARI INSTITUTE OF TECHNOLOGY & MANAGEMENT (BITM) CAMPUS, BALLARI-HOSPET ROAD, NEAR ALLIPUR, BALLARI-583104': 'Kishkinda University',
    'Kalpatharu Institute of Technology, Tiptur B.H.ROAD,HASSAN CIRCLE,TIPTUR-572202': 'Kalpatharu Institute of Technology',
    'Sri Basaveswara Institute of Technology, Madenur Gate, B.H. Road, Tiptur': 'Sri Basaveswara Institute of Technology',
    'Shridevi Institute of Engineering & Technology, Tumkur MARELENAHALLY,SIRA ROAD ,TUMKUR': 'Shridevi Institute of Engineering & Technology',
    'Siddaganga Institute of Technology,Tumkur': 'Siddaganga Institute of Technology',
    'Sri Siddhartha Institute of Technology MARALUR ,TUMKUR': 'Sri Siddhartha Institute of Technology',
    'SRI SIDDHARTHA SCHOOL OF ENGINEERING, TUMKUR NH-4, KESARAMADU POST, KYATHASANDRA,TUMKUR': 'Sri Siddhartha School of Engineering',
    'Channabasaveshwara Institute of Technology, Gubbi Herur, Tumkur NH 206 (BH ROAD), GUBBI, TUMKUR. KARNATAKA': 'Channabasaveshwara Institute of Technology',
    'Akshaya Institute of Technology, Lingapura, Tumkur Dist. LINGAPURA OBLAPURA POST,TUMKUR KORATAGERE ROAD,TUMKUR': 'Akshaya Institute of Technology',
    'C Byre Gowda Institute of Technology, Thoradevandahalli Village, Kolar SRINIVASPURA ROAD, THORADEVANDAHALLI, KOLAR-563101': 'C Byre Gowda Institute of Technology',
    'S J C Institute of Technology, Chickkaballapur(AUTONOMOUS) P B NO. 20, B B ROAD, CHICKBALLAPUR, CHICKBALLAPUR (DISTRICT)': 'S J C Institute of Technology',
    'Dr.T.Thimmaiah Institute of Technology, KGF 00RGAUM POST,KOLAR GOLD FIELDS - 563 120': 'Dr. T. Thimmaiah Institute of Technology',
    'K L S Viswanathrao Deshpande Institute of Technology, Haliyal UDYOG VIDYA NAGAR,DANDELI ROAD, HALIYAL-581329': 'K L S Viswanathrao Deshpande Institute of Technology',
    'Tontadarya College of Engineering, Gadag MUNDARAGI ROAD ,GADAG - 582101': 'Tontadarya College of Engineering',
    "R.T.E. Soceity's Rural Engineering College, Hulkoti HULKOTI -582 205,GADAG": "R T E Society's Rural Engineering College",
    'Angadi Institute of Technology and Management , Savgoan Rd., Belgaum SAVAGAON ROAD BELGAUM': 'Angadi Institute of Technology and Management',
    'Jain College of Engineering,Machche, Belgaum T.S. NAGAR,HUNCHANHATTI ROAD,MACCHE, BELGAUM- 590014.': 'Jain College of Engineering',
    'Jain College Of Engineering and Research, Belgaum SHREYAS CTS 1598 ADAJCENT TO TO FOUNDRY CLUSTER NEAR DUTCH INDUSTRIAL ESTATE CHANNAMMA NAGAR ANAGOL BELGAVI': 'Jain College of Engineering and Research',
    "BLDEA's VP. Dr.P.G. Hallakatti College of Engineering & Technology, Bijapur ASHRAM ROAD,BIJAPUR - 586103": "BLDEA's V P Dr P G Hallakatti College of Engineering & Technology",
    'Maratha Mandal Engineering College, Belgaum RS. NO.104, HALBHAVI VILLAGE, POST: NEW VANTMURI, VIA KAKATI, TALUKA & DIST. BELGAUM-591113': 'Maratha Mandal Engineering College',
    'S.S.E.T.S.S.G.Balekundri Institute of Technology, Shivabasavanagar, Belgaum SHIVABASAVA NAGAR, BELGAUM - 590010': 'S S E T S S G Balekundri Institute of Technology',
    'Secab Institute of Engineering & Technology, Bijapur 424, NAURASPUR, BAGALKOT ROAD, HUDCO CROSS,BIJAPUR-586101': 'Secab Institute of Engineering & Technology',
    "B V V Sangha`s Basaveshwara Engineering College , Bagalkote (AUTONOMOUS) NIJALINGAPPA ROAD,BAGALKOT 587102": 'B V V Sangha Basaveshwara Engineering College',
    'Biluru Gurubasava Mahaswamiji Institute of Technology,Mudhol,Bagalkote. MANTUR ROAD, MUDHOL,BAGALKOTE': 'Biluru Gurubasava Mahaswamiji Institute of Technology',
    'Anjuman Institute of Technology & Management, Bhatkal ANJUMANABAD,BHATKAL,(U K DIST), PIN 581320': 'Anjuman Institute of Technology & Management',
    'Hira Sugar Institute of Technology, Nidasoshi NIDAOSHI,TALUKA: HUKERI,DISTRICT: BELGAUM': 'Hira Sugar Institute of Technology',
    "V S M`s Somashekhar R Kothiwale Institute of Technology,Nippani, Belgaum. BASAV-VIDYA NAGAR,BELGAUM": "V S M's Somashekhar R Kothiwale Institute of Technology",
    'Jain College of Engineering and Technology,Hubballi 404/4, 404/5, SAI NAGAR, UNKAL, HUBLI': 'Jain College of Engineering and Technology',
    'JAIN COLLEGE OF ENGINEERING AND TECHNOLOGY,HUBBALLI': 'Jain College of Engineering and Technology',
    'Adhichunchanagiri Institute of Technology, Chickamagalur ADICHUNCHANAGIRI EXTENTION,JYOTHINAGAR,CHIKMAGALUR - 577102': 'Adhichunchanagiri Institute of Technology',
    'Coorg Institute of Technology, Kunda, Ponnampet POST BOX NO.11,COORG': 'Coorg Institute of Technology',
    'Smt. Kamala & Sri Venkappa M.Agadi College of Engineering & Technology, Gadag SAVANUR ROAD, GULAGANJIKOPPA,LAXMESHWAR DIST: GADAG': 'Smt. Kamala & Sri Venkappa M Agadi College of Engineering & Technology',
    'Sri Taralabalu Jagadguru Institute of Technology, Ranebennur P B ROAD (NH4),RANEBENNUR-581 115,HAVERI DISTRICT': 'Sri Taralabalu Jagadguru Institute of Technology',
    'AGM Rural Engineering College, Varur, Hubli NH-4, P.B. ROAD OPP VRL HEAD OFFICE VARUR.,DHARWAD': 'AGM Rural Engineering College',
    'K C T Engineering College, Gulbarga KCT CAMPUS QAMAR UL ISLAM COLONY,ROZA B,GULBARGA 585104': 'K C T Engineering College',
    "K L E S's K L E College of Engineering & Technology, Chikkodi, Belgaum Dist. BANANTIKODI ROAD CHIKODI,BELGAUM -591 201,": "K L E S K L E College of Engineering & Technology",
    'Khaja Bandanawaz University KBN University Street, Khaja Colony, Kalaburagi, Karnataka 585104': 'Khaja Bandanawaz University',
    'Sharanbasava University (Exclusively for Women) (Formerly Goduati Engineering College For Women) SHARNBASVESHWARA INSTITUTIONS CAMPUS , KALABURAGI ( GULBARGA) -585103 , KARNATAKA': 'Sharanbasava University (Exclusively for Women)',
    'Sharanbasava University (Formerly Appa Institute of Engineering and Tech) SHARNBASVESHWARA INSTITUTIONS CAMPUS , KALABURAGI ( GULBARGA) -585103 , KARNATAKA': 'Sharanbasava University',
    'Shetty Institute of Technology,Gulbarga 5TH KM RAJAPUR SHAHABAD ROAD GULBARGA': 'Shetty Institute of Technology',
    'Lingarajappa Engineering College, Bidar SURVEY NO. 25/1, GORNALLI ,BIDAR': 'Lingarajappa Engineering College',
    'Bheemanna Khandre Institute of Technology, Bhalki BHALKI-HUMNABAD ROAD,BHALKI- 585328,DIST. BIDAR': 'Bheemanna Khandre Institute of Technology',
    'Gurunanak Dev Engineering College, Bidar MAILOOR ROAD,BIDAR 585403': 'Gurunanak Dev Engineering College',
    'Veerappa Nisty Engineering College, Shorapur, Yadigir POST:HASNAPUR, TQ:SHORAPUR, DIST: YADGIR': 'Veerappa Nisty Engineering College',
    'Government Engineering College, Raichur RAICHUR-HYDERABAD ROAD YERAMARUS,RAICHUR': 'Government Engineering College, Raichur',
    'Government Engineering College, Talakal, Koppal TALAKAL,TALUKA: YALABURGA, DIST: KOPPAL, KARNATAKA, PIN 583238': 'Government Engineering College, Talakal',
    'Government Engineering College, Viprasinagar, Gangavathi VIPRASAI NAGAR,GANGAVATHI,KOPPAL': 'Government Engineering College, Gangavathi',
    'Government Engineering College, Hoovina Hadagali HUVINAHADAGALI 583219,DIST: VIJAYANAGARA': 'Government Engineering College, Hoovina Hadagali',
    'Government Engineering College, Ramanagaram DODDAMANNUGUDDE, RAMANAGAR, RAMANAGAR DISTRICT - 571 511': 'Government Engineering College, Ramanagaram',
    'Government Engineering College, Haveri DEVAGIRI,HAVERI': 'Government Engineering College, Haveri',
    'Government Engineering College, Karwar MAJALI VILLAGE, KARWAR TALUK, UTTARA KANNADA DIST, KARNATAKA -PIN 581345': 'Government Engineering College, Karwar',
    'Government Engineering College, Kushalanagar MEDIKERI ROAD,MADAPATNA,KUSHALANAGAR': 'Government Engineering College, Kushalanagar',
    'Government Engineering College, Chamarajanagara GOVERNMENT POLYTECHNIC CAMPUS,NANJANGUD ROAD,CHAMARAJANAGARA': 'Government Engineering College, Chamarajanagara',
    'Government Engineering College, Challakere, Chitradurga BALLARI ROAD CHALLAKERE,CHITRADURGA': 'Government Engineering College, Challakere',
    'Government Engineering College,Mailur,Bidar GANDHI NAGAR ITI CAMPUS MAILUR,BIDAR': 'Government Engineering College, Mailur',
    'Government Engineering College,Naragund,Gadag RON ROAD,SOMAPURA,NARAGUND 582207': 'Government Engineering College, Naragund',
    'K R PET KRISHNA, GOVERNMENT ENGINEERING COLLEGES, K R PET, MANDYA KRISHNARAJPET, MANDYA DIST. - 571 426,': 'Government Engineering College, K R Pet',
    'GOVERNMENT ENGINEERING COLLEGE, ARASIKERE ARASIKERE, HASSAN DISTRICT': 'Government Engineering College, Arasikere',
    'CONSTITUENT COLLEGE OF VTU, CHINTAMANI CHIKABALLAPURA DIST CHIKBALLAPURA': 'Constituent College of VTU, Chintamani',
    'VTU CONSTITUENT ENGINEERING COLLEGE - GOKAK': 'VTU Constituent Engineering College, Gokak',
    'VISVESVARAYA TECHNOLOGICAL UNIVERSITY, BELAGAVI CAMPUS, BELAGAVI JNANA SANGAMA, MACHHE, BELAGAVI': 'Visvesvaraya Technological University (Belagavi Campus)',
    'VISVESVARAYA TECHNOLOGICAL UNIVERSITY, BELAGAVI, VIAT, MUDDENAHALLI CAMPUS, CHIKKABALLAPUR Muddenahalli campus,Chickkaballapur': 'Visvesvaraya Technological University (Muddenahalli Campus)',
    'Visvesvaraya Technological University,VTU,CPGS,Kalburgi. VTU CPGS & REGIONAL OFFICE, KUSNOOR ROAD, KALABURAGI': 'Visvesvaraya Technological University (Kalaburagi)',
    'Visvesvaraya Technological University,VTU,CPGS,Mysuru. Mysuru.': 'Visvesvaraya Technological University (Mysuru)',
    'Sri Venkateshwara College of Engineering, Bangalore VIDYANAGAR, BENGALURU INTERNATIONAL AIRPORT ROAD,BETTAHALASUR POST,BENGALURU NORTH,BENGALURU - 562 157': 'Sri Venkateshwara College of Engineering',
    'ADITYA COLLEGE OF ENGINEERING AND TECHNOLOGY SURVEY NO 2/8 2/9 AND 2/2, KAMAKSHIPURA, SONNENAHALLI PANCHAYATH, HESARAGHATTA HOBLI, YELAHANKA': 'Aditya College of Engineering and Technology',
    'AKASH INTITUTE OF ENGINEERING AND TECHNOLOGY PRASANNAHALLI MAIN RAOD, AKKUPETE VILLAGE, KASABA HOBLI': 'Akash Institute of Engineering and Technology',
    'ANUVARTIK MIRJI BHARATESH INSTITUTE OF TECHNOLOGY BELAGAVI': 'Anuvartik Mirji Bharatesh Institute of Technology',
    'BASAV ENGINEERING SCHOOL OF TECHNOLOGY, VIJAYAPURA': 'Basav Engineering School of Technology',
    'HARSHA INSTITUTE OF TECHNOLOGY VARADANAYAKANA HALLY VILLAGE, NELAMANGALA TALUK, BENALURU RURAL': 'Harsha Institute of Technology',
    'NEW EBENEZER INSTITUTE OF TECHNOLOGY HENNUR BAGALUR MAIN ROAD, KOTHNUR POST, BENGALURU URBAN': 'New Ebenezer Institute of Technology',
    'RAI TECHNOLOGICAL UNIVERSITY 11TH MILE GALLU, DODDABALLAPUR - NELAMANGALA ROAD, MALLOHALLI VILLAGE KADANUR, DODDABALLAPUR TALUK, BENGALURU': 'Rai Technological University',
    'RATHINAM INSTITUTE OF TECHNOLOGY DODDAKAMMANAHALLY VILLAGE, BEGUR HOBLI, BENGALURU SOUTH - 560076': 'Rathinam Institute of Technology',
    'SESHADRIPURAM INSTITUTE OF TECHNOLOGY MYSURU MYSURU': 'Seshadripuram Institute of Technology',
    'DR H N NATIONAL COLLEGE OF ENGINEERING BENGALURU': 'Dr H N National College of Engineering',
    'GANDHI INSTITUTE OF TECHNOLOGY AND MANAGEMENT GITAM OFF CAMPUS BENGALURU NH 207, Nagadenehalli Doddaballapur taluk. Bengaluru-561203': 'Gandhi Institute of Technology and Management (GITAM)',
    'THE CHANAKYA UNIVERSITY NO 29 HARALURU, DEVANAHALLI TALUK (NEAR BENGALURU INTERNATIONLA AIRPORT)': 'The Chanakya University',
    'City Engineering College,Doddakalisandra, Bangalore South DODDAKALLASANDRA, KANAKAPURA ROAD, BANGALORE.': 'City Engineering College',
    'S J M Institute of Technology, Chitradurga P B NO 73, NH -4 BYE PASS, CHITRADURGA': 'S J M Institute of Technology',
    'BMS College of Architecture,Basavanagudi,Bull Temple Road,,BMSCE Campus,Bangalore': 'BMS College of Architecture',
    'SJB School of Achitecture & Planning, Uthrahalli Road, Kengeri, Bangalore 67 BGS Health and Education City, Uttarahalli Road, Kengeri, Bangalore': 'SJB School of Architecture & Planning',
    'Vivekananada Institute of Technology, Kengeri, Bangalore GUDIMAVU VILLAGE ,KENGERI(HOBLI)NEAR KUMBALAGODU,BANGALORE': 'Vivekananda Institute of Technology',
    'University of Mysuru B.N.BAHADUR INSTITUTE OF MANAGEMENT SCIENCES, (DOS IN BUSSINESS ADMINISTRATION)UNIVERSITY OF MYSORE, MANASAGANGOTRI,HUNSUR ROAD,MYSORE.': 'University of Mysuru (B N Bahadur Institute of Management Sciences)',
}

COURSE_CLEAN = {
    'ELECTRONICS AND COMMUNICATIO N ENGG': 'ELECTRONICS AND COMMUNICATION ENGINEERING',
    'ELECTRONICS AND COMMUNICATIO N ENGG (VLSI DESIGN AND TECHNOLOGY)': 'ELECTRONICS AND COMMUNICATION ENGINEERING (VLSI DESIGN AND TECHNOLOGY)',
    'ELECTRONICS AND COMMUNICATIO N (ADVANCED COMMUNICATIO N TECHNOLOGY)': 'ELECTRONICS AND COMMUNICATION ENGINEERING (ADVANCED COMMUNICATION TECHNOLOGY)',
    'ELECTRONICS & COMMUNICATIO N ENGINEERING(I NDUSTRIAL INTEGTATED)': 'ELECTRONICS AND COMMUNICATION ENGINEERING (INDUSTRIAL INTEGRATED)',
    'ELECTRONICS AND INSTRUMENTATI ON ENGINEERING': 'ELECTRONICS AND INSTRUMENTATION ENGINEERING',
    'ELECTRONICS & INSTRUMENTATI ON ENGINEERING': 'ELECTRONICS AND INSTRUMENTATION ENGINEERING',
    'ELECTRONICS AND TELECOMMUNIC ATION ENGINEERING': 'ELECTRONICS AND TELECOMMUNICATION ENGINEERING',
    'COMPUTER AND COMMUNICATIO N ENGINEERING': 'COMPUTER AND COMMUNICATION ENGINEERING',
    'COMMUNICATIO N DESIGN': 'COMMUNICATION DESIGN',
    'CIVIL ENVIRONMENTA L ENGINEERING': 'CIVIL ENVIRONMENTAL ENGINEERING',
    'ENVIRONMENTA L ENGINEERING': 'ENVIRONMENTAL ENGINEERING',
    'MECHANICAL AND SMART MANUFACTURIN G': 'MECHANICAL AND SMART MANUFACTURING',
    'BIO- TECHNOLOGY': 'BIO-TECHNOLOGY',
    'Artificial Intelligence Engg': 'ARTIFICIAL INTELLIGENCE ENGINEERING',
    'COMPUTER SCIENCE AND ENGG (ARTIFICIAL INTELLIGENCE)': 'COMPUTER SCIENCE AND ENGINEERING (ARTIFICIAL INTELLIGENCE)',
    'COMPUTER SCIENCE AND ENGG(ARTIFICIA L INTELLIGENCE AND MACHINE LEARNING)': 'COMPUTER SCIENCE AND ENGINEERING (ARTIFICIAL INTELLIGENCE AND MACHINE LEARNING)',
    'COMPUTER SCIENCE AND ENGG(INTERNE T OF THINGS & CYBER SECURITY INCLUDING BLOCK CHAIN TECH)': 'COMPUTER SCIENCE AND ENGINEERING (INTERNET OF THINGS & CYBER SECURITY INCLUDING BLOCK CHAIN TECHNOLOGY)',
    'COMPUTER SCIENCE AND ENGG(INTERNE T OF THINGS)': 'COMPUTER SCIENCE AND ENGINEERING (INTERNET OF THINGS)',
    'COMPUTER SCIENCE AND ENGINEERING (AIML)': 'COMPUTER SCIENCE AND ENGINEERING (ARTIFICIAL INTELLIGENCE AND MACHINE LEARNING)',
    'COMPUTER SCIENCE AND ENGINEERING(A RTIFICAL INTELLIGENCE & DATA SCIENCE)': 'COMPUTER SCIENCE AND ENGINEERING (ARTIFICIAL INTELLIGENCE AND DATA SCIENCE)',
    'COMPUTER SCIENCE AND ENGINEERING(D ATA SCIENCE)': 'COMPUTER SCIENCE AND ENGINEERING (DATA SCIENCE)',
    'COMPUTER SCIENCE AND TECHNOLOGY(E XCLUSIVELY FOR DIFFERENTLY ABLED)': 'COMPUTER SCIENCE AND TECHNOLOGY (EXCLUSIVELY FOR DIFFERENTLY ABLED)',
    'ELECTRONICS ENGINEERING(V LSI DESIGN & TECHNOLOGY)': 'ELECTRONICS ENGINEERING (VLSI DESIGN AND TECHNOLOGY)',
    'BTECH IN ELECTRONICS ENGINEERING(V LSI DESIGN & TECHNOLOGY)': 'ELECTRONICS ENGINEERING (VLSI DESIGN AND TECHNOLOGY)',
    'BTECH IN INFORMATION TECHNOLOGY AUGMENTED REALITY AND VIRUTAL REALITY(AR/VR)': 'INFORMATION TECHNOLOGY (AUGMENTED REALITY AND VIRTUAL REALITY)',
    'BTECH IN INFORMATION TECHNOLOGY DATA ANALYTICS': 'INFORMATION TECHNOLOGY (DATA ANALYTICS)',
    'BTECH IN MECHANICAL AND SMART MANUFACTURIN G': 'MECHANICAL AND SMART MANUFACTURING',
    'BTECH IN PHARMACEUTIC AL ENGINEERING': 'PHARMACEUTICAL ENGINEERING',
    'BTECH IN COMPUTER SCIENCE AND BUSINESS SYSTEMS': 'COMPUTER SCIENCE AND BUSINESS SYSTEMS',
    'BTECH IN COMPUTER SCIENCE AND DESIGN': 'COMPUTER SCIENCE AND DESIGN',
    'AERO SPACE ENGINEERING': 'AEROSPACE ENGINEERING',
    'B TECH IN AERO SPACE ENGINEERING': 'AEROSPACE ENGINEERING',
    'B TECH IN AERONAUTICAL ENGINEERING': 'AERONAUTICAL ENGINEERING',
    'B TECH IN AGRICULTURAL ENGINEERING': 'AGRICULTURE ENGINEERING',
    'B TECH IN ARTIFICIAL INTELLIGENCE AND DATA SCIENCE': 'ARTIFICIAL INTELLIGENCE AND DATA SCIENCE',
    'B TECH IN ARTIFICIAL INTELLIGENCE AND MACHINE LEARNING': 'ARTIFICIAL INTELLIGENCE AND MACHINE LEARNING',
    'B TECH IN BIO- TECHNOLOGY': 'BIO-TECHNOLOGY',
    'B TECH IN CIVIL ENGINEERING': 'CIVIL ENGINEERING',
    'B TECH IN COMPUTER ENGINEERING': 'COMPUTER ENGINEERING',
    'B TECH IN COMPUTER SCIENCE': 'COMPUTER SCIENCE (B.TECH)',
    'B TECH IN COMPUTER SCIENCE AND ENGINEERING': 'COMPUTER SCIENCE AND ENGINEERING',
    'B TECH IN COMPUTER SCIENCE AND INFORMATION TECHNOLOGY': 'COMPUTER SCIENCE AND INFORMATION TECHNOLOGY',
    'B TECH IN COMPUTER SCIENCE AND TECHNOLOGY': 'COMPUTER SCIENCE AND TECHNOLOGY',
    'B TECH IN COMPUTER SCIENCE (CLOUD COMPUTING)': 'COMPUTER SCIENCE AND ENGINEERING (CLOUD COMPUTING)',
    'B TECH IN COMPUTER SCIENCE (CYBER SECURITY)': 'COMPUTER SCIENCE AND ENGINEERING (CYBER SECURITY)',
    'B TECH IN COMPUTER SCIENCE (DATA SCIENCE)': 'COMPUTER SCIENCE AND ENGINEERING (DATA SCIENCE)',
    'B TECH IN COMPUTER SCIENCE & ENGINEERING (ARTIFICAL INTELLIGENCE & MACHINE LEARNING)': 'COMPUTER SCIENCE AND ENGINEERING (ARTIFICIAL INTELLIGENCE AND MACHINE LEARNING)',
    'B TECH IN COMPUTER SCIENCE & ENGG (ARTIFICIAL INTELLIGENCE AND FUTURE TECHNOLOGIES )': 'COMPUTER SCIENCE AND ENGINEERING (ARTIFICIAL INTELLIGENCE AND FUTURE TECHNOLOGIES)',
    'B TECH IN COMPUTER SCIENCE AND ENGINEERING(A RTIFICIAL INTELLIGENCE AND DATA SCIENCE)': 'COMPUTER SCIENCE AND ENGINEERING (ARTIFICIAL INTELLIGENCE AND DATA SCIENCE)',
    'B TECH IN COMPUTER SCIENCE AND ENGINEERING(B LOCK CHAIN)': 'COMPUTER SCIENCE AND ENGINEERING (BLOCKCHAIN)',
    'B TECH IN COMPUTER SCIENCE AND ENGINEERING(C YBER SECURITY)': 'COMPUTER SCIENCE AND ENGINEERING (CYBER SECURITY)',
    'B TECH IN COMPUTER SCIENCE AND ENGINEERING(D ATA SCIENCE)': 'COMPUTER SCIENCE AND ENGINEERING (DATA SCIENCE)',
    'B TECH IN COMPUTER SCIENCE AND ENGINEERING(I OT)': 'COMPUTER SCIENCE AND ENGINEERING (INTERNET OF THINGS)',
    'B TECH IN COMPUTER SCIENCE AND ENGINEERING(I OT INCLUDING BLOCK CHAIN)': 'COMPUTER SCIENCE AND ENGINEERING (INTERNET OF THINGS INCLUDING BLOCKCHAIN)',
    'B TECH IN COMPUTER SCIENCE AND TECHNOLOGY(B IG DATA)': 'COMPUTER SCIENCE AND TECHNOLOGY (BIG DATA)',
    'B TECH IN COMPUTER SCIENCE AND TECHNOLOGY(D EV OPS)': 'COMPUTER SCIENCE AND TECHNOLOGY (DEVOPS)',
    'B TECH IN ELECTRICAL & ELECTRONICS ENGINEERING': 'ELECTRICAL AND ELECTRONICS ENGINEERING',
    'B TECH IN ELECTRONICS & COMMUNICATIO N ENGINEERING': 'ELECTRONICS AND COMMUNICATION ENGINEERING',
    'B TECH IN ELECTRONICS & COMPUTER ENGINEERING': 'ELECTRONICS AND COMPUTER ENGINEERING',
    'B TECH IN ENERGY ENGINEERING': 'ENERGY ENGINEERING',
    'B TECH IN INFORMATION SCIENCE & TECHNOLOGY': 'INFORMATION SCIENCE AND TECHNOLOGY',
    'B TECH IN INFORMATION SCIENCE ENGINEERING': 'INFORMATION SCIENCE AND ENGINEERING',
    'B TECH IN INFORMATION TECHNOLOGY': 'INFORMATION TECHNOLOGY',
    'B TECH IN MATHAMATICS AND COMPUTING': 'MATHEMATICS AND COMPUTING',
    'B TECH IN MECHANICAL ENGINEERING': 'MECHANICAL ENGINEERING',
    'B TECH IN MECHATRONICS ENGINEERING': 'MECHATRONICS ENGINEERING',
    'B TECH IN PETROLEUM ENGINEERING': 'PETROLEUM ENGINEERING',
    'B TECH IN ROBOTIC ENGINEERING': 'ROBOTICS ENGINEERING',
    'B TECH IN ROBOTICS AND AUTOMATION': 'ROBOTICS AND AUTOMATION',
    'B TECH IN ROBOTICS ENGINEERING': 'ROBOTICS ENGINEERING',
    'B Tech in Computer Science (Information Security)': 'COMPUTER SCIENCE AND ENGINEERING (INFORMATION SECURITY)',
    'B Tech in Computer Science(AI &ML)': 'COMPUTER SCIENCE AND ENGINEERING (ARTIFICIAL INTELLIGENCE AND MACHINE LEARNING)',
    'B Tech in ROBOTICS AND ARTIFICIAL INTELLIGENCE': 'ROBOTICS AND ARTIFICIAL INTELLIGENCE',
    'B.Plan': 'BACHELOR OF PLANNING',
    'B.TECH IN CIVIL CONSTRUCTION AND SUSTAINABILITY ENGINEERING': 'CIVIL CONSTRUCTION AND SUSTAINABILITY ENGINEERING',
    'B.TECH IN COMPUTER ENGINEERING(S OFTWARE PRODUCT DEVELOPMENT)': 'COMPUTER ENGINEERING (SOFTWARE PRODUCT DEVELOPMENT)',
    'B.TECH IN COMPUTER SCIENCE AND ARTIFICIAL INTELLIGENCE': 'COMPUTER SCIENCE AND ENGINEERING (ARTIFICIAL INTELLIGENCE)',
    'B.TECH IN COMPUTER SCIENCE AND ENGG (ROBOTICS)': 'COMPUTER SCIENCE AND ENGINEERING (ROBOTICS)',
    'B.TECH IN COMPUTER SICENCE AND ENGG (DATA ANALYTICS)': 'COMPUTER SCIENCE AND ENGINEERING (DATA ANALYTICS)',
    'B.TECH IN Computer Science and Medical Engineering': 'COMPUTER SCIENCE AND MEDICAL ENGINEERING',
    'B.TECH IN ELECTRICAL ENGINEERING AND COMPUTER SCIENCE': 'ELECTRICAL ENGINEERING AND COMPUTER SCIENCE',
    'B.TECH IN ELECTRONICS ENGINEERING': 'ELECTRONICS ENGINEERING',
    'B.TECH IN ELECTRONICS ENGINEERING (VLSI AND EMBEDDED SYSTEM)': 'ELECTRONICS ENGINEERING (VLSI AND EMBEDDED SYSTEMS)',
    'B.TECH IN EMBEDDED SYSTEM AND VLSI': 'EMBEDDED SYSTEMS AND VLSI',
    'B.TECH IN MECHANICAL AND AEROSPACE ENGINEERING': 'MECHANICAL AND AEROSPACE ENGINEERING',
    'B.Tech In BIOTECHNOLOG Y & BIO- ENGINEERING': 'BIOTECHNOLOGY AND BIO-ENGINEERING',
    'B.Tech in COMPUTER SCIENCE & ENGG (Business Systems)': 'COMPUTER SCIENCE AND ENGINEERING (BUSINESS SYSTEMS)',
    'B.Tech in Computer Science (Internet of Things)': 'COMPUTER SCIENCE AND ENGINEERING (INTERNET OF THINGS)',
    'B.Tech in Computer Science and Engineering(Clou d Computing)': 'COMPUTER SCIENCE AND ENGINEERING (CLOUD COMPUTING)',
    'B.Tech in Computer Science and Engineering(Dev Ops)': 'COMPUTER SCIENCE AND ENGINEERING (DEVOPS)',
    'B.Tech in Computer Science and Engineering(Full Stack Development)': 'COMPUTER SCIENCE AND ENGINEERING (FULL STACK DEVELOPMENT)',
    'B.Tech in Electrical and Electronics Engineering (Electrical Vehicle Technology)': 'ELECTRICAL AND ELECTRONICS ENGINEERING (ELECTRIC VEHICLE TECHNOLOGY)',
    'B.Tech in VLSI': 'ELECTRONICS ENGINEERING (VLSI)',
    'B TECH (HONS) COMPUTER SCIENCE AND ENGINEERING(D ATA SCIENCE)': 'COMPUTER SCIENCE AND ENGINEERING (DATA SCIENCE) HONOURS',
    'ELECTRICAL & ELECTRONICS ENGINEERING': 'ELECTRICAL AND ELECTRONICS ENGINEERING',
    'ELECTRICAL & COMPUTER ENGINEERING': 'ELECTRICAL AND COMPUTER ENGINEERING',
    'ELECTRONICS & COMPUTER ENGINEERING': 'ELECTRONICS AND COMPUTER ENGINEERING',
    'ELECTRONICS & COMPUTER SCIENCE': 'ELECTRONICS AND COMPUTER SCIENCE',
    'INDUSTRIAL ENGINEERING & MANAGEMENT': 'INDUSTRIAL ENGINEERING AND MANAGEMENT',
    'CERAMICS & CEMENT ENGINEERING': 'CERAMICS AND CEMENT ENGINEERING',
    'INDUSTRIAL & PRODUCTION ENGINEERING': 'INDUSTRIAL AND PRODUCTION ENGINEERING',
    'CONSTRUCTION TECHNOLOGY AND MGMT': 'CONSTRUCTION TECHNOLOGY AND MANAGEMENT',
    'DATA SCIENCES': 'DATA SCIENCE',
    'MECHATRONICS': 'MECHATRONICS ENGINEERING',
    'BIO-MEDICAL ENGINEERING': 'BIOMEDICAL ENGINEERING',
    'BIOMEDICAL AND ROBOTIC ENGINEERING': 'BIOMEDICAL AND ROBOTICS ENGINEERING',
    'POLYMER SCIENCE & TECHNOLOGY': 'POLYMER SCIENCE AND TECHNOLOGY',
    'COMPUTER SCIENCE & TECHNOLOGY': 'COMPUTER SCIENCE AND TECHNOLOGY',
    'COMPUTER SCIENCE AND ENGINEERING(DATA SCIENCE)': 'COMPUTER SCIENCE AND ENGINEERING (DATA SCIENCE)',
}

df['College_Name'] = df['College_Name'].map(lambda x: COLLEGE_CLEAN.get(x, x))
df['Course_Name'] = df['Course_Name'].map(lambda x: COURSE_CLEAN.get(x, x))
df[CATEGORIES] = df[CATEGORIES].where(df[CATEGORIES].notna(), other=None)

# Validation
assert not df['College_Code'].isnull().any()
for col in CATEGORIES:
    orig = original_values[col].dropna().tolist()
    new = df[col].dropna().tolist()
    assert orig == new, f"Values changed in {col}"

duplicates = df.duplicated(subset=['College_Code','Course_Name']).sum()
total_colleges = df['College_Code'].nunique()
total_courses = df['Course_Name'].nunique()
total_records = len(df)
missing_counts = df[CATEGORIES].isnull().sum().sum()

# --- Styling helper using row iteration (avoid named tuple column issues) ---
def style_worksheet(ws, df_out, col_widths):
    header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill  = PatternFill('solid', start_color='1F4E79')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font    = Font(name='Arial', size=9)
    alt_fill     = PatternFill('solid', start_color='EBF3FB')
    white_fill   = PatternFill('solid', start_color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align   = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC')
    )
    cols = df_out.columns.tolist()
    ws.append(cols)
    for ci in range(1, len(cols)+1):
        c = ws.cell(1, ci)
        c.font = header_font; c.fill = header_fill
        c.alignment = header_align; c.border = thin
    ws.row_dimensions[1].height = 36

    TEXT_COLS = {'College_Name', 'Course_Name'}
    for ri, row in enumerate(df_out.values.tolist(), 2):
        fill = alt_fill if ri % 2 == 0 else white_fill
        for ci, (col, val) in enumerate(zip(cols, row), 1):
            if isinstance(val, float) and np.isnan(val):
                val = None
            cell = ws.cell(ri, ci, value=val)
            cell.font = data_font; cell.fill = fill; cell.border = thin
            cell.alignment = left_align if col in TEXT_COLS else center_align

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

# WIDE output
wb_wide = openpyxl.Workbook()
ws_wide = wb_wide.active
ws_wide.title = "Cut-Off Ranks (Wide)"
wide_widths = {'A': 13, 'B': 48, 'C': 42}
for i in range(4, len(df.columns)+1):
    wide_widths[get_column_letter(i)] = 8
style_worksheet(ws_wide, df, wide_widths)
wb_wide.save('/home/icarus/Desktop/scripts/outputs/cleaned_wide.xlsx')
print("Saved: cleaned_wide.xlsx")

# LONG output
id_cols = ['College_Code', 'College_Name', 'Course_Name']
df_long = df.melt(id_vars=id_cols, value_vars=CATEGORIES, var_name='Category', value_name='Cutoff')
df_long = df_long.dropna(subset=['Cutoff']).reset_index(drop=True)
df_long['Cutoff'] = pd.to_numeric(df_long['Cutoff'], errors='coerce')
df_long = df_long.sort_values(['College_Code','Course_Name','Category']).reset_index(drop=True)

wb_long = openpyxl.Workbook()
ws_long = wb_long.active
ws_long.title = "Cut-Off Ranks (Long)"
long_widths = {'A': 13, 'B': 48, 'C': 42, 'D': 8, 'E': 12}
style_worksheet(ws_long, df_long, long_widths)
wb_long.save('/home/icarus/Desktop/scripts/outputs/cleaned_long.xlsx')
print(f"Saved: cleaned_long.xlsx  ({len(df_long)} rows)")

print(f"""
╔══════════════════════════════════════════════════╗
║         VALIDATION SUMMARY REPORT               ║
╠══════════════════════════════════════════════════╣
║  Total colleges (unique):      {total_colleges:<18}║
║  Total branches (unique):      {total_courses:<18}║
║  Total records (wide):         {total_records:<18}║
║  Total records (long):         {len(df_long):<18}║
║  Duplicate records:            {duplicates:<18}║
║  Missing cutoff values (wide): {missing_counts:<18}║
║  Missing college codes:        {df['College_Code'].isnull().sum():<18}║
║  Cutoff values unmodified:     {'YES':<18}║
║  All 28 category cols present: {'YES':<18}║
╚══════════════════════════════════════════════════╝""")
